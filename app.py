import io
import os
import re
import traceback
import openpyxl
import streamlit as st
import pandas as pd

from imputar_core import procesar, aplicar
from comprobantes_helper import cargar_indice
from mep_helper import cargar_mep


@st.cache_data(ttl=3600)
def cargar_mep_cacheado():
    return cargar_mep()

st.set_page_config(page_title="Imputar Cuotas", page_icon="📊", layout="wide")

CACHE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'datos', 'comprobantes_cache.json')

# ── Helpers ──────────────────────────────────────────────────────────────────

def leer_hojas(file_bytes):
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    except Exception:
        return []

def es_hoja_pesos(nombre):
    return bool(re.match(r'^S\s*\d+', nombre, re.IGNORECASE))

def es_hoja_usd(nombre):
    return bool(re.match(r'^USD\s*\d+', nombre, re.IGNORECASE))

def parse_cuota_override(texto):
    resultado = {}
    if not texto.strip():
        return resultado
    for linea in texto.strip().splitlines():
        linea = linea.strip()
        if not linea or linea.startswith('#'):
            continue
        partes = re.split(r'[:\s,]+', linea)
        if len(partes) >= 2:
            cuit = re.sub(r'[^0-9]', '', partes[0])
            try:
                cuota = int(partes[1])
                if cuit:
                    resultado[cuit] = cuota
            except ValueError:
                pass
    return resultado

def fmt_monto(val, es_usd):
    if val is None:
        return ''
    if es_usd:
        return f'U$D {val:,.2f}'
    return f'${val:,.0f}'

def fmt_dif(val, es_usd):
    if val is None:
        return ''
    signo = '+' if val >= 0 else ''
    if es_usd:
        return f'{signo}{val:,.2f}'
    return f'{signo}{val:,.0f}'

def restaurar_cache_formulas(orig_bytes, mod_bytes):
    """Devuelve `mod_bytes` con cada celda-fórmula reemplazada por su valor
    cacheado del ORIGINAL.

    Al procesar varias hojas en cadena, `aplicar()` guarda deudores con openpyxl
    y eso BORRA el valor cacheado de las fórmulas (los 'teorico' son fórmulas).
    Si la siguiente hoja leyera ese deudores, vería teórico = None y saltearía
    imputaciones. Como los teóricos son fijos del mes (no cambian al imputar),
    restauramos el valor original. Esta copia es SOLO para lectura de la próxima
    pasada; el archivo final se regenera desde los bytes originales al confirmar.
    """
    wb_cache = openpyxl.load_workbook(io.BytesIO(orig_bytes), data_only=True)
    wb_mod = openpyxl.load_workbook(io.BytesIO(mod_bytes))
    for ws in wb_mod.worksheets:
        if ws.title not in wb_cache.sheetnames:
            continue
        ws_cache = wb_cache[ws.title]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cached = ws_cache[cell.coordinate].value
                    if cached is not None:
                        cell.value = cached
    out = io.BytesIO()
    wb_mod.save(out)
    wb_mod.close()
    wb_cache.close()
    return out.getvalue()

# ── UI ───────────────────────────────────────────────────────────────────────

st.title('📊 Imputar Cuotas')

try:
    col1, col2 = st.columns(2)
    with col1:
        imp_file = st.file_uploader('Archivo de imputaciones (.xlsx)', type='xlsx', key='imp_upload')
    with col2:
        deu_file = st.file_uploader('Archivo de deudores (.xlsx)', type='xlsx', key='deu_upload')

    if not imp_file or not deu_file:
        st.info('Subí ambos archivos para continuar.')
        st.stop()

    imp_bytes = imp_file.read()
    deu_bytes = deu_file.read()

    hojas = leer_hojas(imp_bytes)
    if not hojas:
        st.error('No se pudo leer el archivo de imputaciones.')
        st.stop()

    # Todas las hojas de semana (pesos o USD), en el orden del archivo (cronológico)
    hojas_semana = [h for h in hojas if es_hoja_pesos(h) or es_hoja_usd(h)]

    if not hojas_semana:
        st.error('No se encontraron hojas de semanas (formato "S 121" o "USD 5") en el archivo.')
        st.stop()

    def etiqueta_hoja(h):
        return f'{h}  ·  {"USD" if es_hoja_usd(h) else "Pesos"}'

    st.divider()

    seleccion = st.multiselect(
        'Semanas a imputar (podés elegir varias, mezclando pesos y USD)',
        options=hojas_semana,
        default=[hojas_semana[-1]],
        format_func=etiqueta_hoja,
        help='Se procesan en orden cronológico. Si un mismo cliente aparece en '
             'varias semanas, la cuota se encadena correctamente (cada semana ve '
             'lo que escribió la anterior).',
    )

    col_tp, col_tu, col_maxrow = st.columns(3)
    with col_tp:
        tol_pesos = st.number_input('Tolerancia pesos', min_value=0, value=3000,
                                    help='Diferencia máxima aceptable en hojas de pesos')
    with col_tu:
        tol_usd = st.number_input('Tolerancia USD', min_value=0, value=5,
                                  help='Diferencia máxima aceptable en hojas USD')
    with col_maxrow:
        max_row = st.number_input('Filas máx.', min_value=1, value=500,
                                  help='Hasta qué fila procesar en cada hoja de imputaciones')

    with st.expander('Overrides manuales de cuota'):
        st.caption('Una línea por CUIT: `20111111112: 12`')
        override_text = st.text_area('CUOTA_OVERRIDE', value='', height=100, label_visibility='collapsed')

    cuota_override = parse_cuota_override(override_text)

    st.divider()

    # ── Simular ───────────────────────────────────────────────────────────────

    if st.button('🔍 Simular', type='primary', use_container_width=True):
        if not seleccion:
            st.warning('Elegí al menos una semana.')
            st.stop()

        # Procesar en el orden del archivo (cronológico), respetando la selección
        orden = [h for h in hojas_semana if h in seleccion]

        comprobantes_cache = cargar_indice(CACHE_PATH) if os.path.exists(CACHE_PATH) else {}
        mep_rates, mep_origen = ({}, None)
        if any(es_hoja_pesos(h) for h in orden):
            mep_rates, mep_origen = cargar_mep_cacheado()

        passes = []
        # Bytes de LECTURA que se encadenan entre hojas: cada hoja lee lo que la
        # anterior ya escribió (para elegir el próximo lote libre y continuar la
        # cuota), pero con el cache de fórmulas del teórico restaurado.
        read_imp, read_deu = imp_bytes, deu_bytes

        with st.spinner('Procesando...'):
            for i, hoja in enumerate(orden):
                es_usd = es_hoja_usd(hoja)
                tol = int(tol_usd if es_usd else tol_pesos)
                logs = []
                results, pago_menos, pago_mas, ambiguous, sin_fila, usd_en_pesos, mes_info, sheets_cfg = procesar(
                    imp_bytes=read_imp,
                    deu_bytes=read_deu,
                    imp_sheet=hoja,
                    es_usd=es_usd,
                    tolerance=tol,
                    max_row=int(max_row),
                    cuota_override=cuota_override,
                    comprobantes_cache=comprobantes_cache,
                    mep_rates=mep_rates,
                    log_fn=logs.append,
                )
                passes.append({
                    'hoja': hoja,
                    'es_usd': es_usd,
                    'results': results,
                    'pago_menos': pago_menos,
                    'pago_mas': pago_mas,
                    'ambiguous': ambiguous,
                    'sin_fila': sin_fila,
                    'usd_en_pesos': usd_en_pesos,
                    'mes_info': mes_info,
                    'sheets_cfg': sheets_cfg,
                    'logs': logs,
                })
                # Preparar los bytes de lectura para la próxima hoja (si hay).
                if i < len(orden) - 1:
                    read_imp, applied_deu = aplicar(
                        results=results,
                        pago_menos=pago_menos,
                        pago_mas=pago_mas,
                        sin_fila=sin_fila,
                        usd_en_pesos=None,  # se aplican recién al confirmar
                        imp_bytes=read_imp,
                        deu_bytes=read_deu,
                        imp_sheet=hoja,
                        sheets_cfg=sheets_cfg,
                    )
                    # Restaurar el cache de fórmulas (teórico) desde el ORIGINAL:
                    # aplicar() lo borra al guardar y la próxima hoja lo necesita.
                    read_deu = restaurar_cache_formulas(deu_bytes, applied_deu)

        st.session_state['sim'] = {
            'passes': passes,
            'mep_origen': mep_origen,
            'imp_bytes': imp_bytes,
            'deu_bytes': deu_bytes,
            'imp_name': imp_file.name,
            'deu_name': deu_file.name,
        }
        st.session_state.pop('descarga', None)

    # ── Mostrar resultados ────────────────────────────────────────────────────

    sim = st.session_state.get('sim')
    if not sim:
        st.stop()

    passes = sim['passes']

    # Totales combinados
    tot_results = sum(len(p['results']) for p in passes)
    tot_usd     = sum(len(p['usd_en_pesos']) for p in passes)
    tot_pmenos  = sum(len(p['pago_menos']) for p in passes)
    tot_pmas    = sum(len(p['pago_mas']) for p in passes)
    tot_amb     = sum(len(p['ambiguous']) for p in passes)
    total       = tot_results + tot_usd + tot_pmenos + tot_pmas + tot_amb

    hojas_txt = ', '.join(p['hoja'] for p in passes)
    st.caption(f'Semanas procesadas (en orden): **{hojas_txt}**')

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric('Total filas', total)
    c2.metric('✅ Para imputar', tot_results)
    c3.metric('💵 USD en pesos', tot_usd)
    c4.metric('⚠️ Pago menos', tot_pmenos)
    c5.metric('🔺 Pago más', tot_pmas)
    c6.metric('❌ Ambiguos', tot_amb)

    # Aprobaciones de USD-en-pesos: dict hoja -> lista de imp_row aprobados
    usd_aprobados_por_hoja = {}

    for p in passes:
        hoja = p['hoja']
        es_usd_sim = p['es_usd']
        results = p['results']
        pago_menos = p['pago_menos']
        pago_mas = p['pago_mas']
        ambiguous = p['ambiguous']
        usd_en_pesos = p['usd_en_pesos']

        mes_vals = list(p['mes_info'].values())
        mes_label = mes_vals[0] if mes_vals else '?'

        n_hoja = len(results) + len(usd_en_pesos) + len(pago_menos) + len(pago_mas) + len(ambiguous)

        st.divider()
        st.header(f'🗓️ {hoja}  ·  {"USD" if es_usd_sim else "Pesos"}')
        st.caption(f'Mes detectado en deudores: **{mes_label}**  ·  {n_hoja} filas')

        if p['logs']:
            with st.expander(f'Log de resoluciones — {hoja}'):
                st.code('\n'.join(p['logs']))

        if results:
            st.subheader(f'✅ Para imputar ({len(results)})')
            rows_ok = []
            for r in results:
                rows_ok.append({
                    'Fila': r['imp_row'],
                    'Cliente': f"{r['cliente']}{r['lote_str']}",
                    'CUIT': r['cuit'],
                    'Monto real': fmt_monto(r['monto_real'], es_usd_sim),
                    'Teórico': fmt_monto(r['monto_teo'], es_usd_sim),
                    'Diferencia': fmt_dif(r['diferencia'], es_usd_sim),
                    'Cuota': r['cuota'],
                    'Fecha': r['fecha'].strftime('%d/%m/%Y') if r['fecha'] else '?',
                    'Hoja deudores': r['hoja'],
                })
            st.dataframe(pd.DataFrame(rows_ok), use_container_width=True, hide_index=True)

        if usd_en_pesos:
            st.subheader(f'💵 Clientes USD fijo que pagaron en pesos ({len(usd_en_pesos)})')
            origen = sim.get('mep_origen')
            st.caption(
                f'Cotización dólar MEP venta del día de la transferencia (fuente: misma serie que Ámbito Financiero'
                f'{", " + origen if origen else ""}). Tildá **Imputar** en los que quieras escribir: '
                'en USD fijo va el monto en pesos como "pago real", más N° de cuota y fecha.'
            )
            rows_usd = []
            for e in usd_en_pesos:
                rows_usd.append({
                    'Imputar': False,
                    'Fila': e['imp_row'],
                    'Cliente': str(e['cliente']),
                    'CUIT': e['cuit'],
                    'Fecha': e['fecha'].strftime('%d/%m/%Y') if e['fecha'] else '?',
                    'Pagó en $': fmt_monto(e['monto_pesos'], False),
                    'MEP venta': f"${e['mep']:,.2f}" if e['mep'] else 'sin dato',
                    'Equiv. U$D': fmt_monto(e['equiv_usd'], True) if e['equiv_usd'] is not None else '?',
                    'Teórico U$D': fmt_monto(e['teo_usd'], True) if e['teo_usd'] is not None else '?',
                    'Dif. U$D': fmt_dif(e['dif_usd'], True) if e['dif_usd'] is not None else '?',
                    'Cuota': e['cuota'],
                })
            df_usd = pd.DataFrame(rows_usd)
            edited_usd = st.data_editor(
                df_usd,
                use_container_width=True,
                hide_index=True,
                disabled=[c for c in df_usd.columns if c != 'Imputar'],
                key=f"usd_editor_{hoja}",
            )
            usd_aprobados_por_hoja[hoja] = [
                int(f) for f, ok in zip(edited_usd['Fila'], edited_usd['Imputar']) if ok
            ]

        if pago_menos:
            st.subheader(f'⚠️ Pago menos ({len(pago_menos)}) — se escribirá "PAGO MENOS" en col H')
            rows_pm = []
            for pm in pago_menos:
                rows_pm.append({
                    'Fila': pm['row'],
                    'Cliente': pm['cliente'],
                    'CUIT': pm['cuit'],
                    'Transferido': fmt_monto(pm['transferido'], es_usd_sim),
                    'Teórico': fmt_monto(pm['teorico'], es_usd_sim),
                    'Diferencia': fmt_monto(pm['diferencia'], es_usd_sim),
                })
            st.dataframe(pd.DataFrame(rows_pm), use_container_width=True, hide_index=True)

        if pago_mas:
            st.subheader(f'🔺 Pago más ({len(pago_mas)}) — se escribirá "PAGO MAS" en col H')
            rows_pmas = []
            for pm in pago_mas:
                rows_pmas.append({
                    'Fila': pm['row'],
                    'Cliente': pm['cliente'],
                    'CUIT': pm['cuit'],
                    'Transferido': fmt_monto(pm['transferido'], es_usd_sim),
                    'Teórico': fmt_monto(pm['teorico'], es_usd_sim),
                    'Diferencia': fmt_dif(pm['diferencia'], es_usd_sim),
                })
            st.dataframe(pd.DataFrame(rows_pmas), use_container_width=True, hide_index=True)

        if ambiguous:
            st.subheader(f'❌ Casos ambiguos ({len(ambiguous)}) — revisar manualmente')
            rows_amb = []
            for a in ambiguous:
                cliente = a.get('cliente', a.get('concepto', ''))
                monto = a.get('monto', '')
                monto_str = fmt_monto(monto, es_usd_sim) if isinstance(monto, (int, float)) else str(monto) if monto else ''
                matches_str = str(a.get('matches', '')) if a.get('matches') else ''
                rows_amb.append({
                    'Fila': a.get('row', '?'),
                    'Motivo': a.get('motivo', ''),
                    'Cliente / Concepto': str(cliente)[:60] if cliente else '',
                    'Monto': monto_str,
                    'Candidatos': matches_str[:80] if matches_str else '',
                })
            st.dataframe(pd.DataFrame(rows_amb), use_container_width=True, hide_index=True)

    # ── Confirmar e imputar ───────────────────────────────────────────────────

    hay_algo = tot_results > 0 or tot_pmenos > 0 or tot_usd > 0
    if not hay_algo:
        st.divider()
        st.info('No hay nada para imputar.')
        st.stop()

    tot_usd_aprob = sum(len(v) for v in usd_aprobados_por_hoja.values())

    st.divider()
    aviso = (f'Esto escribirá **{tot_results}** imputaciones y **{tot_pmenos}** "PAGO MENOS" '
             f'en {len(passes)} hoja(s): {hojas_txt}.')
    if tot_usd:
        aviso += f' USD en pesos habilitados: **{tot_usd_aprob}** de {tot_usd}.'
    st.warning(aviso)

    if st.button('✅ Confirmar e Imputar', type='primary', use_container_width=True):
        with st.spinner('Escribiendo archivos...'):
            # Reproducir la cadena desde los bytes originales, ahora aplicando
            # también los USD-en-pesos que el usuario habilitó por hoja.
            work_imp, work_deu = sim['imp_bytes'], sim['deu_bytes']
            for p in passes:
                hoja = p['hoja']
                aprob_filas = usd_aprobados_por_hoja.get(hoja, [])
                usd_aprobados = [e for e in p['usd_en_pesos'] if e['imp_row'] in aprob_filas]
                work_imp, work_deu = aplicar(
                    results=p['results'],
                    pago_menos=p['pago_menos'],
                    pago_mas=p['pago_mas'],
                    sin_fila=p['sin_fila'],
                    usd_en_pesos=usd_aprobados,
                    imp_bytes=work_imp,
                    deu_bytes=work_deu,
                    imp_sheet=hoja,
                    sheets_cfg=p['sheets_cfg'],
                )
            st.session_state['descarga'] = {
                'imp_out': work_imp,
                'deu_out': work_deu,
                'imp_name': sim['imp_name'],
                'deu_name': sim['deu_name'],
            }

    descarga = st.session_state.get('descarga')
    if descarga:
        st.success('Archivos listos. Descargalos y reemplazá los originales en la carpeta compartida.')
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            st.download_button(
                '⬇️ Descargar imputaciones.xlsx',
                data=descarga['imp_out'],
                file_name=descarga['imp_name'],
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )
        with col_d2:
            st.download_button(
                '⬇️ Descargar deudores.xlsx',
                data=descarga['deu_out'],
                file_name=descarga['deu_name'],
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True,
            )

except Exception as e:
    st.error('Error en la aplicación:')
    st.code(traceback.format_exc())
